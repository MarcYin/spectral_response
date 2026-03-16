"""Parser for the official Sentinel-2 SRF workbook."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..ingest import ParsedArtifacts
from ..io import file_sha256
from ..models import SourceManifest

_HEADER_RE = re.compile(r"^(S2[ABC])_SR_AV_(B\d{1,2}A?)$")
_BAND_ORDER = (
    "B01",
    "B02",
    "B03",
    "B04",
    "B05",
    "B06",
    "B07",
    "B08",
    "B8A",
    "B09",
    "B10",
    "B11",
    "B12",
)
_BAND_INDEX = {band_id: index for index, band_id in enumerate(_BAND_ORDER, start=1)}


def parse_s2_srf_xlsx(workbook_path: Path, manifest: SourceManifest) -> ParsedArtifacts:
    """Parse the official Sentinel-2 SRF workbook for a single sensor unit."""

    if not workbook_path.exists():
        raise FileNotFoundError(f"workbook not found: {workbook_path}")

    sensor_code = _sensor_code_from_manifest(manifest.sensor_unit_id)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_name = f"Spectral Responses ({sensor_code})"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"sheet not found in workbook: {sheet_name}")

    curve_sheet = workbook[sheet_name]
    header_row = next(curve_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    band_columns = _extract_band_columns(header_row, sensor_code)

    equivalent_wavelengths = _parse_equivalent_wavelengths(
        workbook["Equivalent Wavelengths"],
        sensor_code,
    )
    bandwidth_rows = _parse_bandwidth_rows(
        workbook["Bandwidth and mid-wavelength"],
        sensor_code,
    )

    curve_rows: list[dict[str, Any]] = []
    band_support: dict[str, dict[str, float | None]] = {
        band_id: {"min": None, "max": None} for band_id in band_columns
    }
    sampling_nm: float | None = None
    previous_wavelength_nm: float | None = None
    wavelength_min_nm: float | None = None
    wavelength_max_nm: float | None = None

    for row in curve_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        wavelength_nm = float(row[0])
        wavelength_min_nm = wavelength_nm if wavelength_min_nm is None else wavelength_min_nm
        wavelength_max_nm = wavelength_nm
        if previous_wavelength_nm is not None and sampling_nm is None:
            sampling_nm = wavelength_nm - previous_wavelength_nm
        previous_wavelength_nm = wavelength_nm

        for band_id, column_index in band_columns.items():
            raw_value = row[column_index] if column_index < len(row) else 0.0
            response = float(raw_value or 0.0)
            curve_rows.append(
                {
                    "sensor_unit_id": manifest.sensor_unit_id,
                    "representation_variant": manifest.representation_variant,
                    "band_id": band_id,
                    "wavelength_nm": wavelength_nm,
                    "response": response,
                    "curve_origin": "canonical_sampled",
                    "realization_id": None,
                    "is_native": True,
                    "is_trimmed": False,
                    "source_id": manifest.source_id,
                }
            )
            if response > 0.0:
                support = band_support[band_id]
                if support["min"] is None:
                    support["min"] = wavelength_nm
                support["max"] = wavelength_nm

    if sampling_nm is None:
        raise ValueError("unable to infer wavelength sampling from the SRF sheet")

    band_rows: list[dict[str, Any]] = []
    band_metrics: dict[str, dict[str, Any]] = {}
    for band_id in _BAND_ORDER:
        if band_id not in band_columns:
            raise ValueError(f"missing band column for {band_id}")
        if band_id not in equivalent_wavelengths:
            raise ValueError(f"missing equivalent wavelength for {band_id}")
        if band_id not in bandwidth_rows:
            raise ValueError(f"missing bandwidth row for {band_id}")

        support = band_support[band_id]
        bandwidth = bandwidth_rows[band_id]
        band_rows.append(
            {
                "sensor_unit_id": manifest.sensor_unit_id,
                "representation_variant": manifest.representation_variant,
                "band_id": band_id,
                "band_index": _BAND_INDEX[band_id],
                "band_name": band_id,
                "center_wavelength_nm": equivalent_wavelengths[band_id],
                "fwhm_nm": bandwidth["bandwidth_nm"],
                "published_shape_type": manifest.canonical.published_shape_type,
                "band_status": "nominal",
                "native_support_min_nm": support["min"],
                "native_support_max_nm": support["max"],
                "native_sampling_nm": sampling_nm,
                "normalization": manifest.canonical.normalization,
                "has_sampled_curve": True,
                "has_band_spec": False,
            }
        )
        band_metrics[band_id] = {
            "equivalent_wavelength_nm": equivalent_wavelengths[band_id],
            "mid_bandwidth_wa_nm": bandwidth["wa_nm"],
            "bandwidth_nm": bandwidth["bandwidth_nm"],
            "native_support_min_nm": support["min"],
            "native_support_max_nm": support["max"],
        }

    metadata = {
        "source_id": manifest.source_id,
        "sensor_unit_id": manifest.sensor_unit_id,
        "representation_variant": manifest.representation_variant,
        "content_kind": manifest.content_kind.value,
        "source_tier": manifest.source_tier.value,
        "source_type": manifest.source_type.value,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "parser": {
            "module": "rsrf.parsers.sentinel2",
            "function": "parse_s2_srf_xlsx",
            "script": manifest.parser.script,
            "entrypoint": manifest.parser.entrypoint,
        },
        "source_artifact": {
            "path": str(workbook_path),
            "sha256": file_sha256(workbook_path),
            "size_bytes": workbook_path.stat().st_size,
        },
        "curve_table": {
            "row_count": len(curve_rows),
            "band_count": len(_BAND_ORDER),
            "wavelength_min_nm": wavelength_min_nm,
            "wavelength_max_nm": wavelength_max_nm,
            "wavelength_sampling_nm": sampling_nm,
        },
        "band_metrics": band_metrics,
        "manifest": manifest.to_dict(),
    }

    return ParsedArtifacts(
        curve_rows=curve_rows,
        band_rows=band_rows,
        metadata=metadata,
    )


def _sensor_code_from_manifest(sensor_unit_id: str) -> str:
    sensor_unit_id = sensor_unit_id.lower()
    if "2a" in sensor_unit_id:
        return "S2A"
    if "2b" in sensor_unit_id:
        return "S2B"
    if "2c" in sensor_unit_id:
        return "S2C"
    raise ValueError(f"unable to infer Sentinel-2 platform from sensor_unit_id: {sensor_unit_id}")


def _extract_band_columns(header_row: tuple[Any, ...], sensor_code: str) -> dict[str, int]:
    band_columns: dict[str, int] = {}
    for column_index, header in enumerate(header_row[1:], start=1):
        if not isinstance(header, str):
            continue
        match = _HEADER_RE.match(header)
        if not match or match.group(1) != sensor_code:
            continue
        band_id = _normalize_band_id(match.group(2))
        band_columns[band_id] = column_index
    if tuple(band_columns.keys()) != _BAND_ORDER:
        missing = [band_id for band_id in _BAND_ORDER if band_id not in band_columns]
        raise ValueError(f"unexpected Sentinel-2 header layout; missing bands: {missing}")
    return band_columns


def _parse_equivalent_wavelengths(worksheet, sensor_code: str) -> dict[str, float]:
    sensor_column_index: int | None = None
    for row in worksheet.iter_rows(min_row=1, max_row=3, values_only=True):
        for index, value in enumerate(row, start=1):
            if value == sensor_code:
                sensor_column_index = index
                break
        if sensor_column_index is not None:
            break
    if sensor_column_index is None:
        raise ValueError(f"unable to locate {sensor_code} column in Equivalent Wavelengths")

    rows: dict[str, float] = {}
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        band_label = row[1]
        value = row[sensor_column_index - 1]
        if band_label is None or value is None:
            continue
        rows[_normalize_band_id(str(band_label))] = float(value)
    return rows


def _parse_bandwidth_rows(worksheet, sensor_code: str) -> dict[str, dict[str, float]]:
    sensor_column_index: int | None = None
    for row in worksheet.iter_rows(min_row=4, max_row=4, values_only=True):
        for index, value in enumerate(row, start=1):
            if value == sensor_code:
                sensor_column_index = index
                break
    if sensor_column_index is None:
        raise ValueError(f"unable to locate {sensor_code} block in Bandwidth and mid-wavelength")

    wa_column_index = sensor_column_index
    bandwidth_column_index = sensor_column_index + 1
    rows: dict[str, dict[str, float]] = {}
    for row in worksheet.iter_rows(min_row=6, values_only=True):
        band_label = row[0]
        wa_value = row[wa_column_index - 1]
        bandwidth_value = row[bandwidth_column_index - 1]
        if band_label is None or wa_value is None or bandwidth_value is None:
            continue
        rows[_normalize_band_id(str(band_label))] = {
            "wa_nm": float(wa_value),
            "bandwidth_nm": float(bandwidth_value),
        }
    return rows


def _normalize_band_id(raw_band_id: str) -> str:
    raw_band_id = raw_band_id.strip().upper()
    if raw_band_id == "B8A":
        return "B8A"
    if not raw_band_id.startswith("B"):
        raise ValueError(f"unexpected Sentinel-2 band label: {raw_band_id}")
    suffix = raw_band_id[1:]
    if not suffix.isdigit():
        raise ValueError(f"unexpected Sentinel-2 band label: {raw_band_id}")
    return f"B{int(suffix):02d}"
