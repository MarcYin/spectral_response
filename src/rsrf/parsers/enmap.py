"""Parser for official EnMAP spectral band workbooks."""

from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..band_specs import default_band_id
from ..ingest import ParsedBandSpecArtifacts
from ..io import file_sha256
from ..models import SourceManifest

_SHEET_ORDER = (
    ("VNIR", 0),
    ("SWIR", 91),
)


def parse_enmap_band_workbook(workbook_path: Path, manifest: SourceManifest) -> ParsedBandSpecArtifacts:
    """Parse the official EnMAP spectral band update workbook."""

    if not workbook_path.exists():
        raise FileNotFoundError(f"EnMAP workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        band_spec_rows: list[dict[str, Any]] = []
        band_rows: list[dict[str, Any]] = []
        for sheet_name, band_offset in _SHEET_ORDER:
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[0] in (None, ""):
                    continue
                local_band_index = int(row[0])
                center_wavelength_nm = float(row[1])
                fwhm_nm = float(row[2])
                overall_band_index = band_offset + local_band_index
                band_id = default_band_id(overall_band_index)
                band_name = f"{sheet_name}{local_band_index:03d}"

                band_spec_rows.append(
                    {
                        "sensor_unit_id": manifest.sensor_unit_id,
                        "representation_variant": manifest.representation_variant,
                        "band_id": band_id,
                        "band_index": overall_band_index,
                        "center_wavelength_nm": center_wavelength_nm,
                        "fwhm_nm": fwhm_nm,
                        "published_shape_type": manifest.canonical.published_shape_type,
                        "shape_param_json": json.dumps({"subsystem": sheet_name}, sort_keys=True),
                        "band_status": "nominal",
                        "is_official": True,
                        "source_id": manifest.source_id,
                    }
                )
                band_rows.append(
                    {
                        "sensor_unit_id": manifest.sensor_unit_id,
                        "representation_variant": manifest.representation_variant,
                        "band_id": band_id,
                        "band_index": overall_band_index,
                        "band_name": band_name,
                        "center_wavelength_nm": center_wavelength_nm,
                        "fwhm_nm": fwhm_nm,
                        "published_shape_type": manifest.canonical.published_shape_type,
                        "band_status": "nominal",
                        "native_support_min_nm": None,
                        "native_support_max_nm": None,
                        "native_sampling_nm": None,
                        "normalization": manifest.canonical.normalization,
                        "has_sampled_curve": False,
                        "has_band_spec": True,
                    }
                )
    finally:
        workbook.close()

    center_wavelengths = [row["center_wavelength_nm"] for row in band_spec_rows]
    monotonic_centers = all(
        current > previous for previous, current in zip(center_wavelengths, center_wavelengths[1:])
    )

    metadata = {
        "source_id": manifest.source_id,
        "sensor_unit_id": manifest.sensor_unit_id,
        "representation_variant": manifest.representation_variant,
        "content_kind": manifest.content_kind.value,
        "source_tier": manifest.source_tier.value,
        "source_type": manifest.source_type.value,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "parser": {
            "module": "rsrf.parsers.enmap",
            "function": "parse_enmap_band_workbook",
            "script": manifest.parser.script,
            "entrypoint": manifest.parser.entrypoint,
        },
        "source_artifact": {
            "path": str(workbook_path),
            "sha256": file_sha256(workbook_path),
            "size_bytes": workbook_path.stat().st_size,
        },
        "band_spec_table": {
            "row_count": len(band_spec_rows),
            "monotonic_centers": monotonic_centers,
            "sheet_order": [sheet_name for sheet_name, _ in _SHEET_ORDER],
        },
        "field_mapping": manifest.band_spec.to_dict(),
        "realization": manifest.curve_realization.to_dict(),
        "manifest": manifest.to_dict(),
    }
    return ParsedBandSpecArtifacts(
        band_spec_rows=band_spec_rows,
        band_rows=band_rows,
        metadata=metadata,
    )
