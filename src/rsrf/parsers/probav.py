"""Parser for official PROBA-V spectral response workbooks."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..models import SourceManifest
from .common import ParsedBandCurve, build_sampled_curve_artifacts

_BAND_LAYOUT = (
    ("BLUE", 1, {"center_camera": 2, "left_camera": 3, "right_camera": 4}),
    ("RED", 6, {"center_camera": 7, "left_camera": 8, "right_camera": 9}),
    ("NIR", 11, {"center_camera": 12, "left_camera": 13, "right_camera": 14}),
    ("SWIR", 16, {"center_camera": 17, "left_camera": 18, "right_camera": 19}),
)


def parse_probav_srf_workbook(workbook_path: Path, manifest: SourceManifest):
    """Parse the official PROBA-V per-camera SRF workbook."""

    if not workbook_path.exists():
        raise FileNotFoundError(f"PROBA-V workbook not found: {workbook_path}")

    camera_key = manifest.representation_variant
    if camera_key not in {"center_camera", "left_camera", "right_camera"}:
        raise ValueError(
            "PROBA-V representation_variant must be one of 'center_camera', 'left_camera', or 'right_camera'"
        )

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        parsed_bands: list[ParsedBandCurve] = []
        for band_index, (band_id, wavelength_column, response_columns) in enumerate(
            _BAND_LAYOUT,
            start=1,
        ):
            response_column = response_columns[camera_key]
            wavelength_nm: list[float] = []
            response: list[float] = []
            for row in range(2, worksheet.max_row + 1):
                wavelength_value = worksheet.cell(row=row, column=wavelength_column).value
                response_value = worksheet.cell(row=row, column=response_column).value
                if wavelength_value in (None, ""):
                    continue
                wavelength_nm.append(float(wavelength_value))
                response.append(0.0 if response_value in (None, "") else float(response_value))
            parsed_bands.append(
                ParsedBandCurve(
                    band_id=band_id,
                    band_index=band_index,
                    band_name=band_id,
                    wavelength_nm=wavelength_nm,
                    response=response,
                )
            )
    finally:
        workbook.close()

    return build_sampled_curve_artifacts(
        manifest,
        workbook_path,
        parsed_bands,
        parser_module="rsrf.parsers.probav",
        parser_function="parse_probav_srf_workbook",
        extra_metadata={
            "probav_srf": {
                "camera_variant": camera_key,
                "sheet_name": worksheet.title,
            }
        },
    )
