from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from rsrf.io import read_json
from rsrf.parsers.sentinel2 import parse_s2_srf_xlsx
from rsrf.validate import parse_manifest_dict

_BANDS = (
    "B01",
    "B02",
    "B03",
    "B04",
    "B05",
    "B06",
    "B07",
    "B08",
    "B8A",
    "B09",
    "B10",
    "B11",
    "B12",
)


def _build_fixture_workbook(path: Path) -> None:
    workbook = Workbook()

    overview = workbook.active
    overview.title = "Overview"
    overview["A1"] = "Overview"

    spectral = workbook.create_sheet("Spectral Responses (S2C)")
    spectral.append(
        ["SR_WL"]
        + [f"S2C_SR_AV_{band_id.replace('B0', 'B').replace('B8A', 'B8A')}" for band_id in _BANDS]
    )
    spectral.append([300] + [0.0] * len(_BANDS))
    spectral.append([301] + [1.0] * len(_BANDS))
    spectral.append([302] + [0.0] * len(_BANDS))

    equivalent = workbook.create_sheet("Equivalent Wavelengths")
    equivalent.append([None] * 8)
    equivalent.append([None, None, "Equivalent wavelength  (nm)", None, None, None, None, None])
    equivalent.append([None, None, "S2A", "S2B", "S2C", None, None, None])
    for index, band_id in enumerate(_BANDS, start=1):
        equivalent.append([None, band_id, 400.0 + index, 401.0 + index, 402.0 + index, None, None, None])

    bandwidth = workbook.create_sheet("Bandwidth and mid-wavelength")
    bandwidth.append([None] * 9)
    bandwidth.append([None] * 9)
    bandwidth.append([None, "Wavelength at mid-bandwidth and bandwidth (nm)", None, None, None, None, None, None, None])
    bandwidth.append([None, "S2A", None, "S2B", None, "S2C", None, None, None])
    bandwidth.append([None, "Wa", "Bandwidth", "Wa", "Bandwidth", "Wa", "Bandwidth", None, None])
    for index, band_id in enumerate(_BANDS, start=1):
        bandwidth.append([band_id, 500.0 + index, 10.0 + index, 501.0 + index, 11.0 + index, 502.0 + index, 12.0 + index, None, None])

    workbook.save(path)


class Sentinel2ParserTests(unittest.TestCase):
    def test_parser_extracts_curves_bands_and_metadata(self) -> None:
        payload = read_json(ROOT / "rsrf_source_manifest_sentinel2c_v2.json")
        manifest = parse_manifest_dict(payload)

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "fixture.xlsx"
            _build_fixture_workbook(workbook_path)
            artifacts = parse_s2_srf_xlsx(workbook_path, manifest)

        self.assertEqual(len(artifacts.curve_rows), 39)
        self.assertEqual(len(artifacts.band_rows), 13)
        self.assertEqual(artifacts.band_rows[0]["band_id"], "B01")
        self.assertEqual(artifacts.band_rows[8]["band_id"], "B8A")
        self.assertEqual(artifacts.metadata["curve_table"]["band_count"], 13)
        self.assertEqual(artifacts.metadata["curve_table"]["wavelength_sampling_nm"], 1.0)


if __name__ == "__main__":
    unittest.main()
