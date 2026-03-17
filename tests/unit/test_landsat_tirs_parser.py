from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from rsrf.io import read_json
from rsrf.manifests import manifest_path
from rsrf.parsers.landsat_tirs import parse_landsat_tirs_workbook
from rsrf.validate import parse_manifest_dict


def _manifest_for(sensor_unit_id: str) -> dict:
    payload = read_json(manifest_path(ROOT, "rsrf_source_manifest_sentinel2c_v2.json"))
    payload["source_id"] = f"{sensor_unit_id}_landsat_tirs_test"
    payload["sensor_unit_id"] = sensor_unit_id
    payload["title"] = f"{sensor_unit_id} Landsat TIRS test source"
    payload["url"] = "https://landsat.usgs.gov/spectral-characteristics-viewer"
    payload["doc_version"] = "test"
    payload["raw_local_path"] = "sources/raw/test.xlsx"
    payload["file_sha256"] = "test"
    payload["parser"]["script"] = "scripts/ingest/ingest_sampled_curve.py"
    payload["parser"]["entrypoint"] = "parse_landsat_tirs_workbook"
    payload["validation"]["plot_overlay_required"] = False
    payload["validation"]["expected_band_count"] = 2
    return payload


def _build_landsat8_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TIRS BA RSR"
    sheet.append(
        [
            "wavelength [um]",
            "TIRS1 10.8um band average",
            "TIRS2 12.0um band average",
        ]
    )
    sheet.append([10.0, 0.0, 0.0])
    sheet.append([10.1, 1.0, 0.5])
    sheet.append([10.2, 0.0, 0.0])
    workbook.create_sheet("Band Summary")
    workbook.save(path)


def _build_landsat9_fixture(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Description"

    band_10 = workbook.create_sheet("TIRS Band 10 BA RSR")
    band_10.append(["wavelength [um]", "Band 10 Band=Average RSR"])
    band_10.append([10.0, 0.0])
    band_10.append([10.1, 1.0])
    band_10.append([10.2, 0.0])

    band_11 = workbook.create_sheet("TIRS Band 11 BA RSR")
    band_11.append(["wavelength [um]", "Band 11 Band-Average RSR"])
    band_11.append([11.0, 0.0])
    band_11.append([11.1, 1.0])
    band_11.append([11.2, 0.0])
    workbook.create_sheet("Band Summary")
    workbook.save(path)


class LandsatTirsParserTests(unittest.TestCase):
    def test_parser_handles_landsat8_tirs_layout(self) -> None:
        manifest = parse_manifest_dict(_manifest_for("landsat-8_tirs"))
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "landsat8.xlsx"
            _build_landsat8_fixture(workbook_path)
            artifacts = parse_landsat_tirs_workbook(workbook_path, manifest)

        self.assertEqual([row["band_id"] for row in artifacts.band_rows], ["B10", "B11"])
        self.assertEqual(artifacts.curve_rows[0]["wavelength_nm"], 10000.0)

    def test_parser_handles_landsat9_tirs_layout(self) -> None:
        manifest = parse_manifest_dict(_manifest_for("landsat-9_tirs2"))
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "landsat9.xlsx"
            _build_landsat9_fixture(workbook_path)
            artifacts = parse_landsat_tirs_workbook(workbook_path, manifest)

        self.assertEqual([row["band_id"] for row in artifacts.band_rows], ["B10", "B11"])
        self.assertEqual(artifacts.curve_rows[-1]["wavelength_nm"], 11200.0)


if __name__ == "__main__":
    unittest.main()
