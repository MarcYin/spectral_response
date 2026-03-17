from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from rsrf.io import read_json
from rsrf.parsers.enmap import parse_enmap_band_workbook
from rsrf.validate import parse_manifest_dict


def _manifest_payload() -> dict:
    payload = read_json(ROOT / "rsrf_source_manifest_hyperspectral_band_spec_example.json")
    payload["source_id"] = "enmap_band_workbook_test"
    payload["sensor_unit_id"] = "enmap_hsi"
    payload["title"] = "EnMAP workbook parser test"
    payload["url"] = "https://www.enmap.org/data_access/"
    payload["doc_version"] = "test"
    payload["raw_local_path"] = "sources/raw/test_enmap.xlsx"
    payload["file_sha256"] = "test"
    payload["parser"]["script"] = "scripts/ingest/ingest_band_spec_table.py"
    payload["parser"]["entrypoint"] = "parse_enmap_band_workbook"
    payload["validation"]["expected_band_count"] = 4
    payload["curve_realization"]["enabled"] = False
    payload["curve_realization"]["output_representation_variant"] = None
    payload["curve_realization"]["profile_type"] = None
    payload["curve_realization"]["approximation"] = False
    payload["curve_realization"]["approximation_reason"] = None
    payload["curve_realization"]["persist_realized_curves"] = False
    payload["curve_realization"]["grid_policy"] = None
    return payload


class EnmapParserTests(unittest.TestCase):
    def test_parse_enmap_band_workbook_combines_vnir_and_swir(self) -> None:
        manifest = parse_manifest_dict(_manifest_payload())
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "fixture.xlsx"
            workbook = Workbook()
            vnir = workbook.active
            vnir.title = "VNIR"
            vnir.append(["BAND #", "CW (nm)", "FWHM (nm)"])
            vnir.append([1, 420.0, 6.5])
            vnir.append([2, 426.0, 6.6])
            swir = workbook.create_sheet("SWIR")
            swir.append(["BAND #", "CW (nm)", "FWHM (nm)"])
            swir.append([1, 900.0, 9.5])
            swir.append([2, 910.0, 9.6])
            workbook.create_sheet("Note")
            workbook.save(workbook_path)

            artifacts = parse_enmap_band_workbook(workbook_path, manifest)

        self.assertEqual(len(artifacts.band_spec_rows), 4)
        self.assertEqual(artifacts.band_spec_rows[0]["band_id"], "B001")
        self.assertEqual(artifacts.band_spec_rows[2]["band_id"], "B092")
        self.assertAlmostEqual(artifacts.band_rows[3]["center_wavelength_nm"], 910.0)


if __name__ == "__main__":
    unittest.main()
