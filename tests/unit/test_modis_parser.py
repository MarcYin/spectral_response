from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from rsrf.io import read_json
from rsrf.parsers.modis import parse_modis_rsr_workbook
from rsrf.validate import parse_manifest_dict


def _manifest_for(sensor_unit_id: str) -> dict:
    payload = read_json(ROOT / "rsrf_source_manifest_sentinel2c_v2.json")
    payload["source_id"] = f"{sensor_unit_id}_modis_test"
    payload["sensor_unit_id"] = sensor_unit_id
    payload["title"] = f"{sensor_unit_id} MODIS test source"
    payload["url"] = "https://mcst.gsfc.nasa.gov/calibration/parameters"
    payload["doc_version"] = "test"
    payload["raw_local_path"] = "sources/raw/test.xlsx"
    payload["file_sha256"] = "test"
    payload["parser"]["script"] = "scripts/ingest/ingest_sampled_curve.py"
    payload["parser"]["entrypoint"] = "parse_modis_rsr_workbook"
    payload["validation"]["plot_overlay_required"] = False
    payload["validation"]["monotonic_centers_required"] = False
    return payload


def _build_aqua_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "READ_ME"
    sheet["A1"] = "fixture"

    band_1 = workbook.create_sheet("band_1")
    band_1.append(["Band", "Wavelength", "InbandPeakMethodRSR", "Method2"])
    band_1.append([1, 400, 0.0, 0.0])
    band_1.append([1, 410, 1.0, 1.0])
    band_1.append([1, 420, 0.0, 0.0])

    band_2 = workbook.create_sheet("band_2")
    band_2.append(["Band", "Wavelength", "InbandPeakMethodRSR", "Method2"])
    band_2.append([2, 500, 0.0, 0.0])
    band_2.append([2, 510, 1.0, 1.0])
    band_2.append([2, 520, 0.0, 0.0])
    workbook.save(path)


def _build_terra_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "modis_pfm_rsp_mid_chnl_sum"
    sheet.append(["Band 1", None, "Band 2", None])
    sheet.append(["Wvln(mm)", "RSR", "Wvln(mm)", "RSR"])
    sheet.append([0.40, 0.0, 0.50, 0.0])
    sheet.append([0.41, 1.0, 0.51, 1.0])
    sheet.append([0.42, 0.0, 0.52, 0.0])
    workbook.save(path)


class ModisParserTests(unittest.TestCase):
    def test_parser_handles_aqua_band_sheets(self) -> None:
        manifest = parse_manifest_dict(_manifest_for("aqua_modis"))
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "aqua.xlsx"
            _build_aqua_fixture(workbook_path)
            artifacts = parse_modis_rsr_workbook(workbook_path, manifest)

        self.assertEqual(len(artifacts.band_rows), 2)
        self.assertEqual(artifacts.band_rows[0]["band_id"], "B1")
        self.assertEqual(artifacts.curve_rows[0]["wavelength_nm"], 400.0)

    def test_parser_handles_terra_paired_columns(self) -> None:
        manifest = parse_manifest_dict(_manifest_for("terra_modis"))
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "terra.xlsx"
            _build_terra_fixture(workbook_path)
            artifacts = parse_modis_rsr_workbook(workbook_path, manifest)

        self.assertEqual(len(artifacts.band_rows), 2)
        self.assertEqual(artifacts.band_rows[1]["band_id"], "B2")
        self.assertEqual(artifacts.curve_rows[0]["wavelength_nm"], 400.0)


if __name__ == "__main__":
    unittest.main()
